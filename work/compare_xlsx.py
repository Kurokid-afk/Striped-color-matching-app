import json
import sys

from openpyxl import load_workbook


def cell_matrix(ws):
    return [
        [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
        for row in range(1, ws.max_row + 1)
    ]


actual_path, expected_path = sys.argv[1:3]
actual = load_workbook(actual_path, data_only=False)
expected = load_workbook(expected_path, data_only=False)

summary = {
    "actual_sheets": actual.sheetnames,
    "expected_sheets": expected.sheetnames,
    "sheets": [],
    "all_values_match": True,
}

for name in expected.sheetnames:
    a = actual[name] if name in actual.sheetnames else None
    e = expected[name]
    if a is None:
        summary["all_values_match"] = False
        summary["sheets"].append({"name": name, "missing": True})
        continue

    a_values = cell_matrix(a)
    e_values = cell_matrix(e)
    diffs = []
    max_rows = max(len(a_values), len(e_values))
    max_cols = max(a.max_column, e.max_column)
    for row in range(max_rows):
        for col in range(max_cols):
            av = a_values[row][col] if row < len(a_values) and col < len(a_values[row]) else None
            ev = e_values[row][col] if row < len(e_values) and col < len(e_values[row]) else None
            if av != ev:
                diffs.append({"row": row + 1, "col": col + 1, "actual": av, "expected": ev})

    if diffs:
        summary["all_values_match"] = False

    summary["sheets"].append({
        "name": name,
        "actual_dimensions": [a.max_row, a.max_column],
        "expected_dimensions": [e.max_row, e.max_column],
        "value_diff_count": len(diffs),
        "first_diffs": diffs[:10],
        "actual_freeze": str(a.freeze_panes) if a.freeze_panes else None,
        "expected_freeze": str(e.freeze_panes) if e.freeze_panes else None,
        "actual_filter": a.auto_filter.ref,
        "expected_filter": e.auto_filter.ref,
    })

print(json.dumps(summary, ensure_ascii=False, indent=2, default=str))
