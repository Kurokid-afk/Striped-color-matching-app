import json
import sys

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


path = sys.argv[1]
wb = load_workbook(path, data_only=True)

for ws in wb.worksheets:
    print(json.dumps({
        "sheet": ws.title,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "merges": [str(item) for item in ws.merged_cells.ranges],
    }, ensure_ascii=True))

    for row in range(1, ws.max_row + 1):
        items = []
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            if isinstance(cell, MergedCell):
                continue
            fill = None
            if cell.fill.fill_type:
                color = cell.fill.fgColor
                fill = color.rgb if color.type == "rgb" else f"{color.type}:{color.indexed or color.theme}"
            if cell.value is not None or fill:
                items.append({
                    "cell": cell.coordinate,
                    "value": cell.value,
                    "fill": fill,
                })
        if items:
            print(json.dumps({"row": row, "items": items}, ensure_ascii=True, default=str))
