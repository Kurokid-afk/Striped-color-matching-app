import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def color_info(color):
    if color is None:
        return None
    return {
        "type": color.type,
        "rgb": color.rgb,
        "indexed": color.indexed,
        "theme": color.theme,
        "tint": color.tint,
    }


def workbook_summary(path_str):
    path = Path(path_str)
    wb = load_workbook(path, data_only=False)
    sheets = []
    for ws in wb.worksheets:
        cells = []
        for row in ws.iter_rows():
            for cell in row:
                has_fill = cell.fill.fill_type is not None
                if cell.value is None and not has_fill:
                    continue
                entry = {"ref": cell.coordinate, "value": cell.value}
                if has_fill:
                    entry["fill"] = {
                        "type": cell.fill.fill_type,
                        "fg": color_info(cell.fill.fgColor),
                        "bg": color_info(cell.fill.bgColor),
                    }
                cells.append(entry)
        sheets.append(
            {
                "title": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "merged": [str(rng) for rng in ws.merged_cells.ranges],
                "cells": cells,
            }
        )
    return {"path": str(path), "sheets": sheets}


if __name__ == "__main__":
    print(json.dumps([workbook_summary(p) for p in sys.argv[1:]], ensure_ascii=False, indent=2, default=str))
