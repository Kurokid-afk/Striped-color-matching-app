import json
import sys

from openpyxl import load_workbook


def color_value(color):
    if color is None:
        return None
    if color.type == "rgb":
        return color.rgb
    if color.type == "theme":
        return f"theme:{color.theme}:{color.tint}"
    return f"{color.type}:{color.indexed}"


def side_style(side):
    return getattr(side, "style", None)


for path in sys.argv[1:]:
    wb = load_workbook(path)
    print(json.dumps({"file": path, "sheets": wb.sheetnames}, ensure_ascii=False))
    for ws in wb.worksheets:
        refs = ["A1", "B1", "A2", "B2"]
        if ws.max_column >= 4:
            refs += ["D1", "D4", "E2"]
        else:
            refs += ["B4", "C2"]
        cells = []
        for ref in refs:
            cell = ws[ref]
            cells.append({
                "ref": ref,
                "style_id": cell.style_id,
                "font": {
                    "name": cell.font.name,
                    "size": cell.font.sz,
                    "bold": cell.font.bold,
                    "color": color_value(cell.font.color),
                },
                "fill": {
                    "type": cell.fill.fill_type,
                    "fg": color_value(cell.fill.fgColor),
                },
                "alignment": {
                    "horizontal": cell.alignment.horizontal,
                    "vertical": cell.alignment.vertical,
                    "wrap": cell.alignment.wrap_text,
                },
                "border": {
                    "left": side_style(cell.border.left),
                    "right": side_style(cell.border.right),
                    "top": side_style(cell.border.top),
                    "bottom": side_style(cell.border.bottom),
                },
            })
        print(json.dumps({
            "sheet": ws.title,
            "widths": {k: v.width for k, v in ws.column_dimensions.items()},
            "default_row_height": ws.sheet_format.defaultRowHeight,
            "custom_row_heights": {
                str(i): ws.row_dimensions[i].height
                for i in range(1, ws.max_row + 1)
                if ws.row_dimensions[i].height is not None
            },
            "freeze": str(ws.freeze_panes) if ws.freeze_panes else None,
            "filter": ws.auto_filter.ref,
            "cells": cells,
        }, ensure_ascii=False, indent=2))
